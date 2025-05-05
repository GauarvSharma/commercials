import streamlit as st
from openpyxl import load_workbook
import tempfile

def process_excel(uploaded_file):
    valid = {"HD", "BS", "AS", "RS", "CR", "GS", "SS", "CD", "TS"}
    wb = load_workbook(uploaded_file)

    main_sheets = {ws.title: ws for ws in wb.worksheets if ws.title in valid and ws.title != "TS"}
    ts_sheet    = wb["TS"] if "TS" in wb.sheetnames else None

    start_cols = {}
    for title, ws in main_sheets.items():
        for idx, cell in enumerate(ws[1]):
            if cell.value and cell.font and cell.font.bold and f"({title})" in str(cell.value):
                start_cols[title] = idx
                break
    if not start_cols:
        st.error("No bold header with (SHEETNAME) found in main sheets.")
        return None

    pointers = {title: 3 for title in main_sheets}
    max_rows = {title: ws.max_row for title, ws in main_sheets.items()}
    output_lines = []
    processed_rs_keys = set()

    while True:
        any_active = False
        for title, ws in main_sheets.items():
            ptr = pointers[title]
            if ptr > max_rows[title]:
                continue
            any_active = True

            start = start_cols[title]
            row = ws[ptr]
            vals = ["" if c.value is None else str(c.value) for c in row[start:]]
            next_ptr = ptr + 1

            if not any(vals):
                pointers[title] = next_ptr
                continue

            if vals and vals[-1] == "1":
                vals.pop()

            if title == "RS":
                key = row[1].value
                if key and key not in processed_rs_keys:
                    group = []
                    last = ptr
                    for rr in range(ptr, max_rows[title] + 1):
                        comp = ws[rr]
                        if comp[1].value == key:
                            v = ["" if c.value is None else str(c.value) for c in comp[start:]]
                            if v and v[-1] == "1":
                                v.pop()
                            if any(v):
                                group.append("|".join(v).strip() + "|")
                                last = rr
                        else:
                            break
                    if group:
                        output_lines.extend(group)
                        processed_rs_keys.add(key)
                        next_ptr = last + 1

            else:
                output_lines.append("|".join(vals).strip() + "|")

            pointers[title] = next_ptr

        if not any_active:
            break

    if ts_sheet:
        ts_start = None
        for idx, cell in enumerate(ts_sheet[1]):
            if cell.value and cell.font and cell.font.bold and "(TS)" in str(cell.value):
                ts_start = idx
                break

        if ts_start is not None:
            for r in range(3, ts_sheet.max_row + 1):
                row = ts_sheet[r]
                vals = ["" if c.value is None else str(c.value) for c in row[ts_start:]]
                if not any(vals):
                    continue
                if vals and vals[-1] == "1":
                    vals.pop()
                output_lines.append("|".join(vals).strip() + "|")

    if output_lines:
        return "\n".join(output_lines)
    else:
        return None

st.title("commercial CIBIL sheet")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx","xlsm","xls"])

if uploaded_file:
    if st.button("Run"):
        with st.spinner("processing..."):
            result = process_excel(uploaded_file)
            if result:
                st.success("Processing Complete.")
                st.download_button("Download Output", result, file_name=" CIBIL_File_output.tap", mime="text/plain")
            else:
                st.warning("No valid data found to process.")




#https://laxmicibils.streamlit.app/                
